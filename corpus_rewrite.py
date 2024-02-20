# update text ids in a directory of corpus VRT files - writes out a full
# copy of each file
#

from pathlib import Path
import re
import logging
from openpyxl import load_workbook
import sys
import csv


LOGLEVEL = logging.INFO
LOGDIR = "./logs"

logger = logging.getLogger(__name__)

logger.setLevel(LOGLEVEL)
logch = logging.StreamHandler()
logch.setLevel(LOGLEVEL)
logger.addHandler(logch)
logfh = logging.FileHandler(Path(LOGDIR) / "corpus_text_id.log")
logfh.setLevel(LOGLEVEL)
logger.addHandler(logfh)

CORPUS_ORIG = "corpus_orig"
CORPUS_REMAPPED = "corpus_reorder"

METADATA_EXCEL = "data/warlaw_metadata_2023.xlsx"
TSV_OUT = "data/warlaw_cqpweb_metadata.tsv"
STATS_OUT = "data/stats.csv"

# COL_NEW_ID = 1 # 'B'
# COL_OLD_ID = 9 # 'J'
TRUNCATE_TO = 40
TEXT_ID_RE = re.compile(r'(<text id=")([^\"]*)(">.*)')

METADATA = [
    "new_id",
    "date",
    "full_title",
    "short_title",
    "icrc_category",
    "historical",
    "old_id",
    "in_force",
    "entry_in_force",
    "lan1",
    "lan2",
    "lan3",
    "lan4",
    "lan5",
    "lan6",
    "lan7",
]

SPREADSHEET = {
    "new_id": 1,
    "date": 2,
    "full_title": 3,
    "short_title": 4,
    "icrc_category": 5,
    "historical": 6,
    "old_id": 9,
    "in_force": 10,
    "entry_in_force": 11,
    "lan1": 12,
    "lan2": 13,
    "lan3": 14,
    "lan4": 15,
    "lan5": 16,
    "lan6": 17,
    "lan7": 18,
}

TEXT_FIELDS = ["full_title", "short_title"]


def sanitise_handle(cell):
    """Sanitise values according to cqpqweb's rule for 'handles'"""
    v = cell.value
    if v is None:
        return "null"
    return re.sub(r"[^A-Za-z0-9_]", "_", str(v))


def sanitise_title(cell):
    """Text fields need to have single quotes"""
    v = cell.value
    if v is None:
        return ""
    if v[0] == "'":
        return v
    return f"'{v}'"


def sanitise_cell(field, cell):
    """Sanitise the contents of a cell based on whether it's text or handle"""
    if field in TEXT_FIELDS:
        return sanitise_title(cell)
    else:
        return sanitise_handle(cell)


def sanitise_row(row):
    """Crosswalk and sanitise a row from the metadata spreadsheet"""
    cleaned = []
    for field in METADATA:
        cell = row[SPREADSHEET[field]]
        cleaned.append(sanitise_cell(field, cell))
    return cleaned


def load_id_mappings(excelfile):
    wb = load_workbook(excelfile)
    ws = wb.active
    mappings = {}
    n = 1
    records = []
    new_id_col = SPREADSHEET["new_id"]
    old_id_col = SPREADSHEET["old_id"]
    for row in ws.iter_rows():
        new_id = sanitise_handle(row[new_id_col])
        old_id = row[old_id_col].value
        if new_id is not None and old_id is not None:
            if new_id != "Text_ID":
                if TRUNCATE_TO:
                    if len(old_id) > TRUNCATE_TO:
                        logger.warn(f"Corpus id {old_id} is > 40 chars")
                    new_id = new_id[:TRUNCATE_TO]
                mappings[old_id] = (new_id, n)
                logger.info(f"map old {old_id} to new {new_id} {n}")
                n += 1
        records.append(sanitise_row(row))
    print(f"Writing tsv metadata to {TSV_OUT}")
    with open(TSV_OUT, "w") as tsvh:
        tsvwriter = csv.writer(tsvh, dialect="excel-tab")
        for rec in records:
            if not rec[0] == "null" and not rec[0] == "Text_ID":
                tsvwriter.writerow(rec)
    return mappings


def convert_ids(mappings, cfh):
    for line in cfh:
        if m := TEXT_ID_RE.match(line):
            prefix = m.group(1)
            old_id = m.group(2)
            suffix = m.group(3)
            if old_id not in mappings:
                logger.warn(f"Warning: id {old_id} in {corpus_file} not in spreadsheet")
                yield line
            else:
                new_id = mappings[old_id]
                new_line = f"{prefix}{new_id}{suffix}\n"
                yield new_line
        else:
            yield line


def convert_file(mappings, orig_file, new_file):
    with open(new_file, "w") as nfh:
        with open(orig_file, "r") as ofh:
            for newline in convert_ids(mappings, ofh):
                nfh.write(newline)


def check_ids(mappings, orig_file):
    n = 0
    old_ids = mappings.keys()
    new_ids = mappings.values()
    with open(orig_file, "r") as ofh:
        for line in ofh:
            if m := TEXT_ID_RE.match(line):
                old_id = m.group(2)
                found_old = old_id in old_ids
                found_new = old_id in new_ids
                print(f"{n},{old_id},{found_old},{found_new}")
            n += 1


def write_temp_text(target_dir, n, new_id, content):
    new_file = target_dir / Path(f"warlaw_{n:03d}_{new_id}")
    logger.info(f"Writing {new_file}")
    with open(new_file, "w") as nfh:
        nfh.write(content)
        # ensure all corpus text files end with a newline
        if not ord(content[-1]) == 10:
            nfh.write("\n")


mappings = load_id_mappings(METADATA_EXCEL)

files = list(Path(CORPUS_ORIG).glob("war*"))
files.sort()

target_dir = Path(CORPUS_REMAPPED)
target_dir.mkdir(parents=True, exist_ok=True)

stats = {}

for corpus_file in files:
    content = ""
    new_id = None
    n = None
    logger.info(f"Reading {corpus_file}")
    with open(corpus_file) as ofh:
        for line in ofh:
            if m := TEXT_ID_RE.match(line):
                prefix = m.group(1)
                old_id = m.group(2)
                suffix = m.group(3)
                if old_id not in mappings:
                    logger.error(
                        f"Warning: id {old_id} in {corpus_file} not in spreadsheet"
                    )
                    sys.exit(-1)
                if new_id is not None:
                    write_temp_text(target_dir, n, new_id, content)
                (new_id, n) = mappings[old_id]
                content = f"{prefix}{new_id}{suffix}\n"
                stats[new_id] = 0
            else:
                content += line
                if not line[0] == "<":
                    stats[new_id] += 1
    if new_id is not None:
        write_temp_text(target_dir, n, new_id, content)

with open(STATS_OUT, "w") as csvh:
    csvwriter = csv.writer(csvh, dialect="excel")
    for new_id, count in stats.items():
        csvwriter.writerow([new_id, count])


#    new_file = target_dir / corpus_file.name
# logger.info(f"{corpus_file} => {new_file}")
# convert_file(mappings, corpus_file, new_file)
