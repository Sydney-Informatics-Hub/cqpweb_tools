# update text ids in a directory of corpus VRT files - writes out a full
# copy of each file
# 

from pathlib import Path
import re
import logging
from openpyxl import load_workbook
import sys

LOGLEVEL = logging.INFO
LOGDIR = "./logs"

logger = logging.getLogger(__name__)

logger.setLevel(LOGLEVEL)
logch = logging.StreamHandler()
logch.setLevel(LOGLEVEL)
logger.addHandler(logch)
logfh = logging.FileHandler(Path(LOGDIR) / "corpus_text_id.log")
logfh.setLevel(LOGLEVEL)
logger.addHandler(logfh)



CORPUS_ORIG = "corpus_orig"
CORPUS_REMAPPED = "corpus_remapped"

METADATA_EXCEL = "List of treaties_CQWEBMetadata_UPDATED2023.xlsx"

COL_NEW_ID = 1 # 'B'
COL_OLD_ID = 6 # 'G'
TRUNCATE_TO = 40
TEXT_ID_RE = re.compile(r'(<text id=")([^\"]*)(">.*)')

def load_id_mappings(excelfile):
    wb = load_workbook(excelfile)
    ws = wb.active
    mappings = {}
    for row in ws.iter_rows():
        new_id = row[COL_NEW_ID].value
        old_id = row[COL_OLD_ID].value
        if not new_id is None and not old_id is None:
            if new_id != 'Text ID':
                if TRUNCATE_TO:
                    old_id = old_id[:TRUNCATE_TO]
                mappings[old_id] = new_id
                logger.info(f"map old {old_id} new {new_id}")
    return mappings

def convert_ids(mappings, cfh):
    for line in cfh:
        if m := TEXT_ID_RE.match(line):
            prefix = m.group(1)
            old_id = m.group(2)
            suffix = m.group(3)
            if not old_id in mappings:
                logger.warn(f"Warning: id {old_id} in {corpus_file} not in spreadsheet")
                yield line
            else:
                new_id = mappings[old_id]
                new_line = f"{prefix}{new_id}{suffix}\n"
                yield new_line
        else:
            yield line


def convert_file(mappings, orig_file, new_file):
    with open(new_file, "w") as nfh:
        with open(orig_file, "r") as ofh:
            for newline in convert_ids(mappings, ofh):
                nfh.write(newline)

def check_ids(mappings, orig_file):
    n = 0;
    print(orig_file)
    old_ids = mappings.keys()
    new_ids = mappings.values()
    with open(orig_file, "r") as ofh:
        for line in ofh:
            if m := TEXT_ID_RE.match(line):
                old_id = m.group(2)
                found_old = old_id in old_ids
                found_new = old_id in new_ids
                print(f"{n},{old_id},{found_old},{found_new}")
            n += 1


mappings = load_id_mappings(METADATA_EXCEL)

files = list(Path(CORPUS_ORIG).glob("war*"))
files.sort()

for corpus_file in files:
    check_ids(mappings, corpus_file)
    new_file = Path(CORPUS_REMAPPED) / corpus_file.name
    logger.info(f"{corpus_file} => {new_file}")
    convert_file(mappings, corpus_file, new_file)
